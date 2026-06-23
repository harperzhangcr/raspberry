from collections import defaultdict
from pathlib import Path
import csv
import re

from openpyxl import load_workbook

WORKBOOK = Path("/Users/chocho/Documents/raspberry/outputs/贷后风险等级卡+处置动作卡0602-2_动作角色编码版_含统一动作.xlsx")
OUT_DIR = Path("/Users/chocho/Documents/raspberry/outputs/action_role_coding_final")
ACTION_CODE_CSV = OUT_DIR / "最终动作编码表.csv"
ACTION_MAPPING_CSV = OUT_DIR / "最终动作映射表.csv"
ROLE_CODE_CSV = OUT_DIR / "最终角色编码表.csv"


def text(value):
    return "" if value is None else str(value).strip().replace("\n", " ")


def compact(value):
    return re.sub(r"\s+", "", text(value))


def split_roles(role):
    role = text(role)
    if not role:
        return []
    return [part.strip() for part in role.replace("/", "；").split("；") if part.strip()]


def unique_join(values):
    return "；".join(sorted({value for value in values if value}))


def find_header(ws):
    for row in range(1, min(ws.max_row, 12) + 1):
        headers = [compact(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if "动作代码" in headers and any(header in headers for header in ("统一动作", "统一处置动作")):
            return row, headers
    return None, []


def find_role_col(headers, unified_col):
    for col in range(unified_col + 1, len(headers) + 1):
        if headers[col - 1] == "角色":
            return col
    for col in range(unified_col - 1, 0, -1):
        if headers[col - 1] == "角色":
            return col
    return None


def original_action(ws, row, original_col):
    value = text(ws.cell(row, original_col).value)
    header = compact(ws.cell(3, original_col).value)
    if header == "角色-动作" and "-" in value:
        return value.split("-", 1)[1].strip()
    return value


def existing_action_order():
    order = {"常规": {}, "异常": {}}
    if not ACTION_CODE_CSV.exists():
        return order
    with ACTION_CODE_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        for idx, row in enumerate(csv.DictReader(f), start=1):
            card_type = row.get("卡片类型", "")
            action = row.get("统一动作", "")
            code = row.get("动作编码", "")
            if card_type not in order or not action:
                continue
            try:
                num = int(code.replace("ACT", ""))
            except ValueError:
                num = 10000 + idx
            order[card_type][action] = (num, idx)
    return order


def existing_role_codes():
    codes = {}
    if not ROLE_CODE_CSV.exists():
        return codes
    with ROLE_CODE_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            codes[row["统一角色"]] = row["角色编码"]
    return codes


def collect(wb):
    records = defaultdict(int)
    rollups = {
        "常规": defaultdict(lambda: {"originals": set(), "count": 0, "roles": set(), "sheets": set()}),
        "异常": defaultdict(lambda: {"originals": set(), "count": 0, "roles": set(), "sheets": set()}),
    }
    role_rollup = defaultdict(lambda: {"count": 0, "sheets": defaultdict(int)})
    locations = []

    for ws in wb.worksheets:
        if "风险等级卡" in ws.title:
            continue
        header_row, headers = find_header(ws)
        if header_row is None:
            continue
        card_type = "异常" if "异常处置动作" in ws.title else "常规"
        for unified_col, header in enumerate(headers, start=1):
            if header not in ("统一动作", "统一处置动作"):
                continue
            code_col = unified_col + 1 if unified_col + 1 <= len(headers) and headers[unified_col] == "动作代码" else None
            if not code_col:
                continue
            role_col = find_role_col(headers, unified_col)
            role_code_col = role_col + 1 if role_col and role_col + 1 <= len(headers) and headers[role_col] == "角色代码" else None
            original_col = unified_col - 1
            locations.append((ws, header_row, original_col, unified_col, code_col, role_col, role_code_col, card_type))

            for row in range(header_row + 1, ws.max_row + 1):
                unified = text(ws.cell(row, unified_col).value)
                if not unified:
                    continue
                original = original_action(ws, row, original_col)
                role = text(ws.cell(row, role_col).value) if role_col else ""
                records[(original, unified, card_type, role, ws.title)] += 1
                item = rollups[card_type][unified]
                item["originals"].add(original)
                item["count"] += 1
                item["sheets"].add(ws.title)
                for role_item in split_roles(role):
                    item["roles"].add(role_item)
                    role_rollup[role_item]["count"] += 1
                    role_rollup[role_item]["sheets"][ws.title] += 1
    return records, rollups, role_rollup, locations


def assign_action_codes(rollups, order):
    codes = {"常规": {}, "异常": {}}
    for card_type, start in (("常规", 1), ("异常", 100)):
        actions = list(rollups[card_type].keys())
        actions.sort(key=lambda name: order[card_type].get(name, (99999, name)))
        for offset, action in enumerate(actions):
            codes[card_type][action] = f"ACT{start + offset:03d}"
    return codes


def assign_role_codes(role_rollup, existing):
    codes = {}
    used = set()
    for role in sorted(role_rollup):
        if role in existing:
            codes[role] = existing[role]
            used.add(existing[role])
    next_num = 1
    for role in sorted(role_rollup):
        if role in codes:
            continue
        while f"ROLE{next_num:03d}" in used:
            next_num += 1
        codes[role] = f"ROLE{next_num:03d}"
        used.add(codes[role])
    return codes


def update_workbook(locations, action_codes, role_codes):
    action_updates = 0
    role_updates = 0
    for ws, header_row, original_col, unified_col, code_col, role_col, role_code_col, card_type in locations:
        for row in range(header_row + 1, ws.max_row + 1):
            unified = text(ws.cell(row, unified_col).value)
            if unified:
                ws.cell(row, code_col).value = action_codes[card_type][unified]
                action_updates += 1
            if role_col and role_code_col:
                role = text(ws.cell(row, role_col).value)
                if role:
                    ws.cell(row, role_code_col).value = unique_join(role_codes.get(role_item, "") for role_item in split_roles(role))
                    role_updates += 1
    return action_updates, role_updates


def write_outputs(records, rollups, role_rollup, action_codes, role_codes):
    with ACTION_MAPPING_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["原始动作", "统一动作", "动作编码", "卡片类型", "角色", "角色编码", "来源sheet", "出现次数"])
        for (original, unified, card_type, role, sheet), count in sorted(records.items(), key=lambda x: (x[0][2], action_codes[x[0][2]][x[0][1]], x[0][0], x[0][3], x[0][4])):
            writer.writerow([
                original,
                unified,
                action_codes[card_type][unified],
                card_type,
                role,
                unique_join(role_codes.get(role_item, "") for role_item in split_roles(role)),
                sheet,
                count,
            ])

    with ACTION_CODE_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["动作编码", "卡片类型", "统一动作", "原始动作数量", "出现次数", "涉及角色", "来源sheet"])
        for card_type in ("常规", "异常"):
            for unified in sorted(rollups[card_type], key=lambda name: action_codes[card_type][name]):
                item = rollups[card_type][unified]
                writer.writerow([
                    action_codes[card_type][unified],
                    card_type,
                    unified,
                    len(item["originals"]),
                    item["count"],
                    unique_join(item["roles"]),
                    unique_join(item["sheets"]),
                ])

    with ROLE_CODE_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["角色编码", "统一角色", "出现次数", "来源sheet"])
        for role in sorted(role_rollup):
            item = role_rollup[role]
            writer.writerow([
                role_codes[role],
                role,
                item["count"],
                "；".join(f"{sheet}({count})" for sheet, count in sorted(item["sheets"].items())),
            ])


def main():
    wb = load_workbook(WORKBOOK)
    order = existing_action_order()
    records, rollups, role_rollup, locations = collect(wb)
    action_codes = assign_action_codes(rollups, order)
    role_codes = assign_role_codes(role_rollup, existing_role_codes())
    action_updates, role_updates = update_workbook(locations, action_codes, role_codes)
    write_outputs(records, rollups, role_rollup, action_codes, role_codes)
    wb.save(WORKBOOK)
    print(f"normal_actions={len(rollups['常规'])}")
    print(f"abnormal_actions={len(rollups['异常'])}")
    print(f"action_updates={action_updates}")
    print(f"role_updates={role_updates}")


if __name__ == "__main__":
    main()
