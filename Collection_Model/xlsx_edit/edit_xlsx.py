from copy import copy
from pathlib import Path
import re
import openpyxl

src = Path('/Users/chocho/Downloads/东风汽金调研交流问题及参会人员v2.xlsx')
out = Path('/Users/chocho/Documents/raspberry/outputs/xlsx_edit/东风汽金调研交流问题及参会人员v2_拆分后.xlsx')

wb = openpyxl.load_workbook(src)
ws = wb.active

start_row, end_row = 51, 58
start_no = ws.cell(start_row, 1).value

numbered_item_re = re.compile(r'(?:^|\n)\s*\d+\s*[\.、]\s*')

def split_items(text):
    if text is None:
        return ['']
    text = str(text).strip()
    matches = list(numbered_item_re.finditer(text))
    if not matches:
        return [text]
    items = []
    for i, m in enumerate(matches):
        item_start = m.end()
        item_end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        item = text[item_start:item_end].strip()
        if item:
            items.append(item)
    return items or [text]

def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)
    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format
    if src_cell.font:
        dst_cell.font = copy(src_cell.font)
    if src_cell.fill:
        dst_cell.fill = copy(src_cell.fill)
    if src_cell.border:
        dst_cell.border = copy(src_cell.border)
    if src_cell.alignment:
        dst_cell.alignment = copy(src_cell.alignment)
    if src_cell.protection:
        dst_cell.protection = copy(src_cell.protection)

original_rows = []
for r in range(start_row, end_row + 1):
    row_values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    row_styles = [copy(ws.cell(r, c)._style) for c in range(1, ws.max_column + 1)]
    row_number_formats = [ws.cell(r, c).number_format for c in range(1, ws.max_column + 1)]
    row_fonts = [copy(ws.cell(r, c).font) for c in range(1, ws.max_column + 1)]
    row_fills = [copy(ws.cell(r, c).fill) for c in range(1, ws.max_column + 1)]
    row_borders = [copy(ws.cell(r, c).border) for c in range(1, ws.max_column + 1)]
    row_alignments = [copy(ws.cell(r, c).alignment) for c in range(1, ws.max_column + 1)]
    row_protections = [copy(ws.cell(r, c).protection) for c in range(1, ws.max_column + 1)]
    original_rows.append({
        'row': r,
        'values': row_values,
        'styles': row_styles,
        'number_formats': row_number_formats,
        'fonts': row_fonts,
        'fills': row_fills,
        'borders': row_borders,
        'alignments': row_alignments,
        'protections': row_protections,
        'height': ws.row_dimensions[r].height,
        'items': split_items(ws.cell(r, 4).value),
    })

extra_rows = sum(len(x['items']) for x in original_rows) - len(original_rows)

# Preserve the left merged section while allowing row insertion and later extending it.
b_value = ws['B37'].value
b_style_source = ws['B37']
for merged in list(ws.merged_cells.ranges):
    if str(merged) == 'B37:B58':
        ws.unmerge_cells(str(merged))
        break

if extra_rows > 0:
    ws.insert_rows(end_row + 1, extra_rows)

current = start_row
next_no = int(start_no) if isinstance(start_no, int) else int(float(start_no))
max_col = ws.max_column
for original in original_rows:
    block_start = current
    for item in original['items']:
        for c in range(1, max_col + 1):
            cell = ws.cell(current, c)
            cell._style = copy(original['styles'][c - 1])
            cell.number_format = original['number_formats'][c - 1]
            cell.font = copy(original['fonts'][c - 1])
            cell.fill = copy(original['fills'][c - 1])
            cell.border = copy(original['borders'][c - 1])
            cell.alignment = copy(original['alignments'][c - 1])
            cell.protection = copy(original['protections'][c - 1])
        ws.row_dimensions[current].height = original['height']
        ws.cell(current, 1).value = next_no
        ws.cell(current, 2).value = None
        ws.cell(current, 3).value = original['values'][2]
        ws.cell(current, 4).value = item
        next_no += 1
        current += 1
    block_end = current - 1
    if block_end > block_start:
        ws.merge_cells(start_row=block_start, start_column=3, end_row=block_end, end_column=3)
        ws.cell(block_start, 3).value = original['values'][2]

new_end = current - 1
ws.merge_cells(start_row=37, start_column=2, end_row=new_end, end_column=2)
ws['B37'].value = b_value
copy_cell_style(b_style_source, ws['B37'])

# Keep D column wrapped for the newly split questions.
for r in range(start_row, new_end + 1):
    ws.cell(r, 4).alignment = copy(ws.cell(r, 4).alignment)
    ws.cell(r, 4).alignment = ws.cell(r, 4).alignment.copy(wrap_text=True, vertical='center')

out.parent.mkdir(parents=True, exist_ok=True)
wb.save(out)
print(out)
print('new_end', new_end, 'rows_created', new_end - start_row + 1)
